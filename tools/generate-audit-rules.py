#!/usr/bin/env python3
"""Generate audit-rules.js from the RGHS audit matrix workbook + rule-curation.json.

The curation overlay (tools/rule-curation.json) is the source of truth for bundle
rules, code lists, name patterns and action classes, because the workbook's
"Package-master reference" column is semi-structured prose. This script:

  1. loads the curation overlay,
  2. extracts the Duplicate MM Codes sheet from the workbook (well-structured),
  3. sanity-checks every curated code against the workbook text (warns on misses),
  4. emits ../audit-rules.js as a UMD module (content script + Node test friendly).

Usage:  python tools/generate-audit-rules.py
"""

import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "RGHS_Healthcare_Coding_Risk_Audit_Matrix.xlsx"
CURATION = Path(__file__).resolve().parent / "rule-curation.json"
OUTPUT = ROOT / "audit-rules.js"


def load_workbook_text(wb):
    """All cell text of the workbook, for code sanity-checking."""
    chunks = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    chunks.append(str(cell))
    return "\n".join(chunks)


def extract_mm_groups(wb):
    ws = wb["Duplicate MM Codes"]
    groups = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name, count, codes, mapping, los, risk = (row + (None,) * 7)[:6]
        if not name or not codes:
            continue
        code_list = [c.strip().upper() for c in str(codes).split(",") if c.strip()]
        if len(code_list) < 2:
            continue
        groups.append({
            "name": str(name).strip(),
            "codes": code_list,
            "risk": (str(risk).strip() if risk else "Medium"),
            "specialtyMapping": str(mapping).strip() if mapping else "",
        })
    return groups


def collect_curated_codes(curation):
    codes = set()

    def take(entity):
        for c in entity.get("codes", []):
            codes.add(c.upper())

    for rule in curation["bundles"]:
        take(rule["bundle"])
        for comp in rule["components"]:
            take(comp)
    for rule in curation["mutuallyExclusive"]:
        for group in rule["groups"]:
            take(group)
    for rule in curation["addOnDependencies"]:
        take(rule["addOn"])
        take(rule["requiresBase"])
    for rule in curation["combinedAvailable"]:
        for group in rule["components"]:
            take(group)
        take(rule["combined"])
    for rule in curation["adjunctClusters"]:
        for member in rule["members"]:
            take(member)
    for rule in curation["reviewTriggers"]:
        take(rule["entity"])
    return codes


def main():
    curation = json.loads(CURATION.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    workbook_text = load_workbook_text(wb).upper()
    missing = sorted(
        code for code in collect_curated_codes(curation)
        if not re.search(r"(?<![\w-])" + re.escape(code) + r"(?![\w])", workbook_text)
    )
    if missing:
        print("WARNING: curated codes not found verbatim in the workbook "
              "(verify against the package master):", file=sys.stderr)
        for code in missing:
            print(f"  - {code}", file=sys.stderr)

    mm_groups = extract_mm_groups(wb)
    print(f"Extracted {len(mm_groups)} duplicate-MM groups from the workbook")

    rules = {
        "schemaVersion": 1,
        "version": datetime.date.today().isoformat(),
        "effectiveDate": datetime.date.today().isoformat(),
        "source": WORKBOOK.name,
        "generatedBy": "tools/generate-audit-rules.py",
        "bundles": curation["bundles"],
        "mutuallyExclusive": curation["mutuallyExclusive"],
        "addOnDependencies": curation["addOnDependencies"],
        "combinedAvailable": curation["combinedAvailable"],
        "adjunctClusters": curation["adjunctClusters"],
        "reviewTriggers": curation["reviewTriggers"],
        "duplicateMMGroups": mm_groups,
        "settings": curation["settings"],
    }

    payload = json.dumps(rules, indent=2, ensure_ascii=False)
    OUTPUT.write_text(
        "// GENERATED FILE - do not edit by hand.\n"
        "// Regenerate with: python tools/generate-audit-rules.py\n"
        f"// Source: {WORKBOOK.name} + tools/rule-curation.json\n"
        "(function (root, factory) {\n"
        "  const api = factory();\n"
        "  if (typeof module === 'object' && module.exports) module.exports = api;\n"
        "  if (root) root.RGHSAuditRules = api;\n"
        "})(typeof globalThis !== 'undefined' ? globalThis : this, function () {\n"
        "  'use strict';\n"
        f"  return {payload};\n"
        "});\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
